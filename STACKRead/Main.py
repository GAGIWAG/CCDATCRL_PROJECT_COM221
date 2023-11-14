import datetime
import tkinter 
import customtkinter
import math
import pandas as pd
import os
from CTkMessagebox import CTkMessagebox
from PIL import Image, ImageTk
import openpyxl
import collections as clc
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog

def side_panel_switch():
    if (math.floor(title_frame.winfo_x()) == -75):
        protract()
        return
    if (math.floor(title_frame.winfo_x()) == 75):   
        retract()
        return
    
def protract():
    title_frame.place(x=title_frame.winfo_x()+15)
    home_frame.place(x=(title_frame.winfo_x())+30) 
    lithit_main_frame.place(x=(title_frame.winfo_x())+35)
    library_main_frame.place(x=(title_frame.winfo_x())+75) 
    add_books_main_frame.place(x=(title_frame.winfo_x()+60)) 
    print(home_frame.winfo_x())
    search_advanceSearch_frame.place(x=(title_frame.winfo_x())+75)

    if math.floor(title_frame.winfo_x()) == 60:
        Main_app.after_cancel(protract)
    else:
        Main_app.after(10,protract)
    
def retract():
    title_frame.place(x=title_frame.winfo_x()-15)
    home_frame.place(x=(title_frame.winfo_x())+45) 
    lithit_main_frame.place(x=(title_frame.winfo_x())+60)
    library_main_frame.place(x=(title_frame.winfo_x())+45) 
    add_books_main_frame.place(x=(title_frame.winfo_x()+80)) 
    print(home_frame.winfo_x())
    search_advanceSearch_frame.place(x=title_frame.winfo_x()+45)

    if  math.floor(title_frame.winfo_x()) == -60:
        Main_app.after_cancel(retract)
        
    else:
        Main_app.after(10,retract)

def show_main_frame_content(text):

    if text == "Home":
        library_frame.pack_forget()
        search_frame.pack_forget()
        lithit_frame.pack_forget()
        add_books_frame.pack_forget()
        open_book_frame.pack_forget()
        lastread_frame.pack_forget()
        home_frame.pack(fill='both', expand=1)
        
        button_home.configure(state="disabled")
        button_lithit.configure(state="normal")
        button_library.configure(state="normal")
        button_search.configure(state="normal")
        button_addbooks.configure(state="normal")
        
        if button_home.cget("state")=="disabled":
            button_home_img_clicked = customtkinter.CTkImage((Image.open("icons\\homeclicked.png")),size=(45,45))
            button_home.configure(image=button_home_img_clicked)
            button_lithit_img = customtkinter.CTkImage((Image.open("icons\\lithitsunclicked.png")),size=(35,45))
            button_lithit.configure(image=button_lithit_img)
            button_library_img = customtkinter.CTkImage((Image.open("icons\\libraryunclicked.png")),size=(45,45))
            button_library.configure(image=button_library_img)
            button_search_img = customtkinter.CTkImage((Image.open("icons\\searchunclicked.png")),size=(45,45))
            button_search.configure(image=button_search_img)
            button_addbooks_img_clicked = customtkinter.CTkImage((Image.open("icons\\disclickedaddbooks.png")),size=(45,45))
            button_addbooks.configure(image=button_addbooks_img_clicked)
        else:
            button_home_img = customtkinter.CTkImage((Image.open("icons\\homeunclicked.png")),size=(45,45))
            button_home.configure(image=button_home_img)

        
    if text == "lithit":
        library_frame.pack_forget()
        search_frame.pack_forget()
        home_frame.pack_forget()
        add_books_frame.pack_forget()
        open_book_frame.pack_forget()
        lastread_frame.pack_forget()
        lithit_frame.pack(fill='both', expand=1)

        
        button_home.configure(state="normal")
        button_lithit.configure(state="disabled")
        button_library.configure(state="normal")
        button_search.configure(state="normal")
        button_addbooks.configure(state="normal")
        
        if button_lithit.cget("state")=="disabled":
            button_home_img = customtkinter.CTkImage((Image.open("icons\\homeunclicked.png")),size=(45,45))
            button_home.configure(image=button_home_img)
            button_lithit_img_clicked = customtkinter.CTkImage((Image.open("icons\\lithitsclicked.png")),size=(35,45))
            button_lithit.configure(image=button_lithit_img_clicked)
            button_library_img = customtkinter.CTkImage((Image.open("icons\\libraryunclicked.png")),size=(45,45))
            button_library.configure(image=button_library_img)
            button_search_img = customtkinter.CTkImage((Image.open("icons\\searchunclicked.png")),size=(45,45))
            button_search.configure(image=button_search_img)
            button_addbooks_img_clicked = customtkinter.CTkImage((Image.open("icons\\disclickedaddbooks.png")),size=(45,45))
            button_addbooks.configure(image=button_addbooks_img_clicked)
        else:
            button_lithit_img = customtkinter.CTkImage((Image.open("icons\\lithitsunclicked.png")),size=(45,45))
            button_lithit.configure(image=button_lithit_img)
        
    if text == "Search":
        library_frame.pack_forget()
        lithit_frame.pack_forget()
        home_frame.pack_forget()
        add_books_frame.pack_forget()
        open_book_frame.pack_forget()
        lastread_frame.pack_forget()
        search_frame.pack(fill='both', expand=1)
        
        button_home.configure(state="normal")
        button_lithit.configure(state="normal")
        button_search.configure(state="disabled")
        button_library.configure(state="normal")
        button_addbooks.configure(state="normal")
        
        if button_search.cget("state")=="disabled":
            button_home_img = customtkinter.CTkImage((Image.open("icons\\homeunclicked.png")),size=(45,45))
            button_home.configure(image=button_home_img)
            button_lithit_img = customtkinter.CTkImage((Image.open("icons\\lithitsunclicked.png")),size=(35,45))
            button_lithit.configure(image=button_lithit_img)
            button_search_img_clicked = customtkinter.CTkImage((Image.open("icons\\searchclicked.png")),size=(45,45))
            button_search.configure(image=button_search_img_clicked)
            button_library_img = customtkinter.CTkImage((Image.open("icons\\libraryunclicked.png")),size=(45,45))
            button_library.configure(image=button_library_img)
            button_addbooks_img_clicked = customtkinter.CTkImage((Image.open("icons\\disclickedaddbooks.png")),size=(45,45))
            button_addbooks.configure(image=button_addbooks_img_clicked)

        else:
            button_search_img = customtkinter.CTkImage((Image.open("icons\\searchunclicked.png")),size=(45,45))
            button_search.configure(image=button_search_img)

    if text == "Library":
        search_frame.pack_forget()
        lithit_frame.pack_forget()
        home_frame.pack_forget()
        add_books_frame.pack_forget()
        open_book_frame.pack_forget()
        lastread_frame.pack_forget()
        library_frame.pack(fill='both', expand=1)

        button_home.configure(state="normal")
        button_lithit.configure(state="normal")
        button_search.configure(state="normal")
        button_library.configure(state="disabled")
        button_addbooks.configure(state="normal")
        
        if button_library.cget("state")=="disabled":
            button_home_img = customtkinter.CTkImage((Image.open("icons\\homeunclicked.png")),size=(45,45))
            button_home.configure(image=button_home_img)
            button_lithit_img = customtkinter.CTkImage((Image.open("icons\\lithitsunclicked.png")),size=(35,45))
            button_lithit.configure(image=button_lithit_img)
            button_search_img_ = customtkinter.CTkImage((Image.open("icons\\searchunclicked.png")),size=(45,45))
            button_search.configure(image=button_search_img_)
            button_library_img_clicked = customtkinter.CTkImage((Image.open("icons\\libraryclicked.png")),size=(45,45))
            button_library.configure(image=button_library_img_clicked)
            button_addbooks_img_clicked = customtkinter.CTkImage((Image.open("icons\\disclickedaddbooks.png")),size=(45,45))
            button_addbooks.configure(image=button_addbooks_img_clicked)

        else:
            button_library_img = customtkinter.CTkImage((Image.open("icons\\libraryunclicked.png")),size=(45,45))
            button_library.configure(image=button_library_img)

    if text == "Add":
        search_frame.pack_forget()
        lithit_frame.pack_forget()
        home_frame.pack_forget()
        library_frame.pack_forget()
        open_book_frame.pack_forget()
        lastread_frame.pack_forget()
        add_books_frame.pack(fill='both', expand=1)
        
        button_home.configure(state="normal")
        button_lithit.configure(state="normal")
        button_search.configure(state="disabled")
        button_library.configure(state="normal")
        button_addbooks.configure(state="disabled")
        
        if button_addbooks.cget("state")=="disabled":
            button_home_img = customtkinter.CTkImage((Image.open("icons\\homeunclicked.png")),size=(45,45))
            button_home.configure(image=button_home_img)
            button_lithit_img = customtkinter.CTkImage((Image.open("icons\\lithitsunclicked.png")),size=(35,45))
            button_lithit.configure(image=button_lithit_img)
            button_search_img_ = customtkinter.CTkImage((Image.open("icons\\searchunclicked.png")),size=(45,45))
            button_search.configure(image=button_search_img_)
            button_library_img = customtkinter.CTkImage((Image.open("icons\\libraryunclicked.png")),size=(45,45))
            button_library.configure(image=button_library_img)
            button_addbooks_img_clicked = customtkinter.CTkImage((Image.open("icons\\clickedaddbooks.png")),size=(45,45))
            button_addbooks.configure(image=button_addbooks_img_clicked)

        else:
            button_addbooks_img = customtkinter.CTkImage((Image.open("icons\\disclickedaddbooks.png")),size=(45,45))
            button_addbooks.configure(image=button_addbooks_img)

def log_in_clock():
    
    rtc = datetime.datetime.now()
    if (rtc.hour < 12):
        if (rtc.hour == 12):
            hr_min_time = "{:02d}:{:02d}:{:02d} AM".format(rtc.hour+12,rtc.minute,rtc.second)
        else:
            hr_min_time = "{:02d}:{:02d}:{:02d} AM".format(rtc.hour,rtc.minute,rtc.second)
    else:
        hr_min_time = "{:02d}:{:02d}:{:02d} PM".format(rtc.hour-12,rtc.minute,rtc.second)
    login_clock_label.configure(text=hr_min_time)
    Main_app.after(1000,log_in_clock)

def logout():
    home_frame.pack_forget()
    library_frame.pack_forget()
    search_frame.pack_forget()
    lithit_frame.pack_forget()
    add_books_frame.pack_forget()
    log_in_frame.pack(fill='both', expand=1) 

def to_log_in():
    sign_in_frame.pack_forget()
    log_in_frame.pack(fill='both', expand=1)  

def login():
    
    username_email = login_username_entry.get()
    password = login_password_entry.get()

    for stored_username, (stored_email, stored_password) in user_credentials.items():
        if username_email == stored_username or username_email == stored_email:
            if password == stored_password:
                msg = CTkMessagebox(title="LogIn", message="Login Success!",icon="warning", option_1="Ok")
                response = msg.get()
                if response == "Ok":
                    log_in_frame.pack_forget() 
                    show_main_frame_content("Home")
                    login_username_entry.delete(0,tkinter.END)
                    login_password_entry.delete(0,tkinter.END)

                else:
                    msg = CTkMessagebox(title="Error Account not found", message="Account Not Found!",icon="warning", option_1="Ok")
 
def agree():
    value = sign_in_checkbox.get()
    if value == "agree":
        terms_n_conditions = True
        return terms_n_conditions 
    if value == "disagree":
        terms_n_conditions = False
        return terms_n_conditions
    
def to_sign_in():
    log_in_frame.pack_forget()
    sign_in_frame.pack(fill='both', expand=1)    

def sign_in():
    username = sign_in_username_entry.get()
    password = sign_in_password_entry.get()
    email = sign_in_email_entry.get()
    
    if username != "" and password != "" and email != "":
        user_credentials[username] = (email, password)
        save_user_credentials()
        to_log_in()
        
    else:
        msg = CTkMessagebox(message="Please fill  up all the forms",icon="warning", option_1="Ok")
            
def acc_reg_success():
    msg = CTkMessagebox(title="Account Registered", message="Account Registered!",icon="check", option_1="Clock-In")
    response = msg.get()
    
    if response=="Clock-In":
        to_log_in()       

def save_user_credentials():
    dff = pd.DataFrame(user_credentials.items(), columns=['Username', 'Email_Password'])
    dff[['Email', 'Password']] = pd.DataFrame(user_credentials.values())
    dff.drop('Email_Password', axis=1, inplace=True)
    dff.to_excel(usercreds_excel_file, index=False)

def add_book():
    title = addbooks_book_detail_input_title_entry.get()
    author = addbooks_book_detail_input_author_name_entry.get()
    genre = addbooks_book_detail_input_genre_entry.get()
    theme = addbooks_book_detail_input_theme_entry.get()
    rating = addbooks_book_detail_rating.get()
    keywords = addbooks_book_detail_input_keywords_entry.get()
    language = addbooks_book_detail_input_language_entry.get()
    awards = addbooks_book_detail_input_language_entry.get()
    pubdate = addbooks_book_detail_input_publish_date_entry.get()

    if title != "" and author != "":
        # Load existing data from excel
        df = pd.read_excel(books_excel_file)

        # Determine the next available Book ID
        next_book_id = len(df) + 1

        # Let the user choose an image file
        image_path = filedialog.askopenfilename(title="Select Book Cover Image", filetypes=[("Image files", "*.png;*.jpg;*.jpeg")])

        # Create a new DataFrame for the book to be added with the determined Book ID
        new_book = pd.DataFrame({'bookID': [next_book_id], 'title': [title], 'author': [author], 'genre': [genre], 'theme': [theme],
                                 'rating': [rating], 'keywords': [keywords], 'language': [language], 'awards': [awards],
                                 'publish_date': [pubdate], 'image_path': [image_path]})

        # Concatenate the new book DataFrame with the existing DataFrame
        df = pd.concat([df, new_book], ignore_index=True)

        # Save the updated data to the excel file, overwriting the previous data
        df.to_excel(books_excel_file, index=False)

        # Save user credentials and other necessary operations
        save_user_credentials()
        save_to_excel()

        msg = CTkMessagebox(message="Book Added Successfully", icon="check", option_1="Ok")
        book_stack()
    else:
        msg = CTkMessagebox(message="Please fill up Title and book Author", icon="warning", option_1="Ok")
        
def upload_cover():
    image_path = filedialog.askopenfilename(title="Select Book Cover Image", filetypes=[("Image files", "*.png;*.jpg;*.jpeg")])

    if image_path:
        # Destroy existing widgets in addbooks_upload_cover_frame
        for widget in addbooks_upload_cover_frame.winfo_children():
            widget.destroy()

        # Display the selected image in addbooks_upload_cover_frame
        cover_image = Image.open(image_path)
        cover_image = cover_image.resize((250, 360))
        cover_image_tk = ImageTk.PhotoImage(cover_image)

        # Create a label to display the image in addbooks_upload_cover_frame
        cover_label = customtkinter.CTkLabel(addbooks_upload_cover_frame, image=cover_image_tk)
        cover_label.image = cover_image_tk
        cover_label.pack(fill='both', expand=True)
        cover_label.update()
    
def select_book():
    title = input("Enter the title of the book you want to select: ")
    
    # Load existing data from excel
    df = pd.read_excel(books_excel_file)

    # If it's not a valid integer, search by title using str.contains for partial and case-insensitive matching
    found_books = df[df['Title'].str.contains(title, case=False)]
    if found_books.empty:
        print("Book not found in the database.")
    else:
        print("Book Details:")

def delete_book():
    title = input("Enter the title of the book to delete: ")

    # Load existing data from excel
    df = pd.read_excel(books_excel_file)

    try:
        # Create a new DataFrame excluding the book to be deleted
        df = df[df['Title'] != title]

        # Save the updated data to the excel file, overwriting the previous data
        df.to_excel(books_excel_file, index=False)
        save_user_credentials()
        print("Book deleted successfully!")
        save_to_excel()
    except KeyError:
        print("Book is not in the library!")
        
def save_to_excel():
    # Read data from excel into dff DataFrame
    df = pd.read_excel(books_excel_file)

    # Save the data to the Excel file
    df.to_excel(books_excel_file, index=False)
    save_user_credentials()
    print("Data saved to Excel and user credentials saved to excel successfully!")

# Sorting algorithm
def quicksort(arr):
    if len(arr) <= 1:   
        return arr
    pivot = arr[len(arr) // 2]
    left = [x for x in arr if x < pivot]
    middle = [x for x in arr if x == pivot]
    right = [x for x in arr if x > pivot] 
    return quicksort(left) + middle + quicksort(right)

# Create a function to display the stack as buttons in a scrollable frame in Lit Hits
def display_lit_hit_data():
    # Sort in descending order
    sorted_df = df.sort_values(by='rating', ascending=False)
    
    book_sorted_info = {
            'title': row[1],
            'author': row[2],
            'rating': row[5],       
        }
    
    # Create buttons for each row and display them in the GUI
    for index,book_sorted_info in sorted_df.iterrows():
        book_text = f"{book_sorted_info['title']} | Rating :{book_sorted_info['rating']}"
        lit_hit_button = customtkinter.CTkButton(lithit_firstShelf_frame,width=900,height=125,hover_color="#241811",fg_color="#2D1E16", text=book_text,font=("Quando",20), command=lambda info=book_sorted_info: display_book_details(info),corner_radius=0)
        lit_hit_button.pack(padx=15, pady=5)
        in_home_button = customtkinter.CTkButton(home_lithit_frame_inner,anchor="c",width=445,height=40,hover_color="#241811",fg_color="#2D1E16", text=book_text,font=("Quando",13), command=lambda info=book_sorted_info: display_book_details(info),corner_radius=0)
        in_home_button.pack(padx=2,pady=5)

# Create a function to display the details of the selected book
def display_book_details(book_info):
    search_frame.pack_forget()
    lithit_frame.pack_forget()
    home_frame.pack_forget()
    library_frame.pack_forget()
    add_books_frame.pack_forget()
    lastread_frame.pack_forget()
    open_book_frame.pack(fill='both', expand=1)
    
    button_home.configure(state="normal")

    book_title_label.configure(text=book_info['title'])
    book_author_label.configure(text=book_info['author'])
    book_genre_placeholder.configure(text=book_info['genre'])
    book_theme_placeholder.configure(text=book_info['theme'])
    book_keywords_placeholder.configure(text=book_info['keywords'])
    book_rating_placeholder.configure(text=book_info['rating'])
    img_path = str(book_info['image_path'])

    # Load the image
    try:
        original_image = Image.open(img_path)
        resized_image = original_image.resize((250, 360))
        tk_image = ImageTk.PhotoImage(resized_image)

        # Display the image on a label
        image_label = tk.Label(opened_book_picture_frame, image=tk_image)
        image_label.image = tk_image  # Keep a reference to the image to prevent it from being garbage collected
        image_label.pack(pady=10)
    except Exception as e:
        print(f"Error loading image: {e}")
    
    last_read_append(book_info)

# Function to add a button to the frame and the stack
def last_read_append(book_info):
    last_read_button = customtkinter.CTkButton(lRead_firstShelf_frame,width=985,hover_color="#B88B68",fg_color="#614D40",height=135, text=f"{book_info['title']} \nby {book_info['author']}",font=("Quando",35), command=lambda info=book_info: display_book_details(info),corner_radius=15)
    last_read_button.pack(pady=2)
    lit_hit_button_stack.append(last_read_button)
    last_home_read_button = customtkinter.CTkButton(home_lastread_frame_inner,width=445,height=45,hover_color="#B88B68",fg_color="#614D40", text=f"{book_info['title']} \nby {book_info['author']}",font=("Quando",13), command=lambda info=book_info: display_book_details(info),corner_radius=15)
    last_home_read_button.pack(pady=2)
    
# Create a function to display the stack as buttons in a scrollable frame in Library
def book_stack():
    # Create buttons for each book in the stack and add them to the frame
    for book_info in book_info_stack:
        book_button = customtkinter.CTkButton(search_bookshelf_frame,width=890,hover_color="#3D291E",fg_color="#614D40",height=55, text=f"{book_info['title']} \nby {book_info['author']}",font=("Quando",20), command=lambda info=book_info: display_book_details(info),corner_radius=15)
        book_button.pack(padx=5,pady=5)     

def open_last_read():
    search_frame.pack_forget()
    lithit_frame.pack_forget()
    home_frame.pack_forget()
    library_frame.pack_forget()
    add_books_frame.pack_forget()
    lastread_frame.pack_forget()
    open_book_frame.pack_forget()
    lastread_frame.pack(fill='both', expand=1)

#\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\FOR DATABASE\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\#

# File paths
books_excel_file = 'spreadsheets\\books.xlsx'
usercreds_excel_file = 'spreadsheets\\usercreds.xlsx'

# Data Frame to store username and password
user_credentials = {}

# Create a linked list to store bookId values
book_id_linked_list = clc.deque()

# Create a stack to store the buttons
lit_hit_button_stack = clc.deque()


# Load existing user credentials from excel
if os.path.exists(usercreds_excel_file):
    dff = pd.read_excel(usercreds_excel_file)
    for index, row in dff.iterrows():
        user_credentials[row['Username']] = (row['Email'], row['Password'])
        
# Load existing data from Excel
df = pd.read_excel(books_excel_file)
dff = pd.read_excel(usercreds_excel_file)

# Create an empty stack to store data
book_info_stack = []

# Load the Excel file
workbook = openpyxl.load_workbook(books_excel_file)

# Choose a specific sheet within the Excel file, e.g., the first sheet
sheet = workbook.active

# Use a variable to skip the first row (headers)
skip_first_row = True

# Iterate through the rows in the sheet and add data to the stack
for row in sheet.iter_rows(values_only=True):

    book_info = {
        'book_id': row[0],
        'title': row[1],
        'author': row[2],
        'genre': row[3],
        'theme' : row[4],
        'rating': row[5],
        'keywords': row[6],
        'image_path': row[10]  
    }
    book_info_stack.append(book_info)

#\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\FOR GUI\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\#

customtkinter.set_appearance_mode("light")
customtkinter.set_default_color_theme("blue")

# Main Window
Main_app  = customtkinter.CTk()
Main_app.title("STACKRead")
# Main_app.iconbitmap(Image.open('icons\stackreadicon.png'))

# Get the screen dimensions
screen_width = 1600
screen_height = 900

# Calculate the desired window size as a fraction of the screen dimensions
window_width = 1360  
window_height = 765     

# Set the window size and position it in the center of the screen
x_position = (screen_width - window_width) // 2
y_position = (screen_height - window_height) // 4.5

# Set the window's geometry
Main_app.geometry(f"{window_width}x{window_height}+{x_position}+{y_position}")
Main_app.resizable(0,0)

#\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\FOR MAIN CONTENT FRAMES\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\#

main_content_frame = customtkinter.CTkFrame(Main_app,width=Main_app.winfo_screenwidth(),height=Main_app.winfo_screenheight(),fg_color="#F2E8BD")
main_content_frame.place(x = 50, y = 0)

home_frame = customtkinter.CTkFrame(main_content_frame,width=Main_app.winfo_screenwidth(),height=Main_app.winfo_screenheight(),fg_color="#F2E8BD")
lithit_frame = customtkinter.CTkFrame(main_content_frame,width=Main_app.winfo_screenwidth(),height=Main_app.winfo_screenheight(),fg_color="#F2E8BD")
lastread_frame = customtkinter.CTkFrame(main_content_frame,width=Main_app.winfo_screenwidth(),height=Main_app.winfo_screenheight(),fg_color="#F2E8BD")
search_frame = customtkinter.CTkFrame(main_content_frame,width=Main_app.winfo_screenwidth(),height=Main_app.winfo_screenheight(),fg_color="#F2E8BD")
library_frame = customtkinter.CTkFrame(main_content_frame,width=Main_app.winfo_screenwidth(),height=Main_app.winfo_screenheight(),fg_color="#F2E8BD")
add_books_frame = customtkinter.CTkFrame(main_content_frame,width=Main_app.winfo_screenwidth(),height=Main_app.winfo_screenheight(),fg_color="#F2E8BD")
open_book_frame = customtkinter.CTkFrame(main_content_frame,width=Main_app.winfo_screenwidth(),height=Main_app.winfo_screenheight(),fg_color="#F2E8BD")

# Side bar Frames
title_frame = customtkinter.CTkFrame(Main_app,width=150,height=Main_app.winfo_screenheight(),fg_color="#C8DF8C",corner_radius=0,border_width=1,border_color="#C8DF8C")
title_frame.place(x = -75, y = 0)

icon_frame = customtkinter.CTkFrame(Main_app,width=76,height=Main_app.winfo_screenheight(),fg_color="#C8DF8C",corner_radius=0,border_width=1,border_color="#C8DF8C")
icon_frame.place(x = 0, y = 0)

# Credential Frames
sign_in_frame = customtkinter.CTkFrame(Main_app,width=Main_app.winfo_screenwidth(),height=Main_app.winfo_screenheight(),fg_color="#F2E8BD")

log_in_frame = customtkinter.CTkFrame(Main_app,width=Main_app.winfo_screenwidth(),height=Main_app.winfo_screenheight(),fg_color="#F2E8BD")
log_in_frame.pack(fill='both', expand=1)    

# Sign in frame
sign_in_label_img = customtkinter.CTkImage(Image.open('icons\\stackreadicon.png'),size=(75,75))
sign_in_label = customtkinter.CTkLabel(sign_in_frame,text="",image=sign_in_label_img,fg_color="transparent")
sign_in_label.place(x=390,y=100)

sign_in_icon_stack_label_img = customtkinter.CTkImage(Image.open('icons\\logo.png'),size=(475,75))
sign_in_icon_stack_label = customtkinter.CTkLabel(sign_in_frame,image=sign_in_icon_stack_label_img,anchor="w",text="",fg_color="transparent",font=("Quando",25))
sign_in_icon_stack_label.place(x=470,y=100)

sign_in_icon_quote_label = customtkinter.CTkLabel(sign_in_frame,anchor="w",text="Your digital book wiki.",fg_color="transparent",font=("Quando",25))
sign_in_icon_quote_label.place(x=640,y=180)

sign_in_credentials_frame = customtkinter.CTkFrame(sign_in_frame,fg_color="#1E1E1E",corner_radius=25,width=725,height=425)
sign_in_credentials_frame.place(x=310,y=220)

sign_in_acc_reg_label = customtkinter.CTkLabel(sign_in_credentials_frame,anchor="center",text_color="white",text="Account Registration",fg_color="transparent",font=("Quando",45))
sign_in_acc_reg_label.place(x=110,y=30)

sign_in_username_entry = customtkinter.CTkEntry(sign_in_credentials_frame,fg_color="white",placeholder_text="Username",font=("Quando",25),corner_radius=25,width=630,height=50)
sign_in_username_entry.place(x=50,y=100)

sign_in_email_entry = customtkinter.CTkEntry(sign_in_credentials_frame,width=630,height=50,fg_color="white",placeholder_text="Email",corner_radius=25,font=("Quando",25))
sign_in_email_entry.place(x=50,y=160)

sign_in_password_entry = customtkinter.CTkEntry(sign_in_credentials_frame,width=630,height=50,fg_color="white",placeholder_text="Password",corner_radius=25,show="●",font=("Quando",25))
sign_in_password_entry.place(x=50,y=220)

terms_n_conditions = False
sign_in_checkbox = customtkinter.CTkCheckBox(sign_in_credentials_frame,command=agree, onvalue='agree',offvalue='disagree',text_color="white",width=170,height=50,text="I have read and understood the terms and conditions",font=("Quando",15))
sign_in_checkbox.place(x=150,y=270)

sign_register_in_button = customtkinter.CTkButton(sign_in_credentials_frame,command=lambda: sign_in(),text="Register",width=230,height=50,fg_color="#C8DF8C",corner_radius=25,text_color="black",font=("Quando",25))
sign_register_in_button.place(x=250,y=340)

sign_in_already_have_an_account_label = customtkinter.CTkLabel(sign_in_frame,anchor="w",text_color="black",text="Already have an account?",fg_color="transparent",font=("Quando",20))
sign_in_already_have_an_account_label.place(x=410,y=655)

clock_in_button = customtkinter.CTkButton(sign_in_frame,command=lambda: to_log_in(),text="Clock-In",width=150,height=25,fg_color="#606060",corner_radius=25,text_color="white",font=("Quando",20))
clock_in_button.place(x=700,y=655)

# Log in Frame Content
login_icon_label_img = customtkinter.CTkImage(Image.open('icons\\stackreadicon.png'),size=(75,75))
login_icon_label = customtkinter.CTkLabel(log_in_frame,text="",image=login_icon_label_img,fg_color="transparent")
login_icon_label.place(x=390,y=100)

login_icon_stack_label_img = customtkinter.CTkImage(Image.open('icons\\logo.png'),size=(475,75))
login_icon_stack_label = customtkinter.CTkLabel(log_in_frame,image=login_icon_stack_label_img,anchor="w",text="",fg_color="transparent",font=("Quando",25))
login_icon_stack_label.place(x=470,y=100)

login_icon_quote_label = customtkinter.CTkLabel(log_in_frame,anchor="w",text="Your digital book wiki.",fg_color="transparent",font=("Quando",25))
login_icon_quote_label.place(x=640,y=180)

login_credentials_frame = customtkinter.CTkFrame(log_in_frame,fg_color="#B88B68",corner_radius=25,width=725,height=525)
login_credentials_frame.place(x=310,y=220)

login_username_entry = customtkinter.CTkEntry(login_credentials_frame,fg_color="white",placeholder_text="Username/Email",font=("Quando",25),corner_radius=25,width=630,height=50)
login_username_entry.place(x=50,y=40)

login_password_entry = customtkinter.CTkEntry(login_credentials_frame,width=630,height=50,fg_color="white",placeholder_text="Password",corner_radius=25,show="●",font=("Quando",25))
login_password_entry.place(x=50,y=100)

login_button = customtkinter.CTkButton(login_credentials_frame,command=lambda: login(),text="Clock-in",width=230,height=50,fg_color="#C8DF8C",corner_radius=25,text_color="black",font=("Quando",25))
login_button.place(x=250,y=170)

login_clock_frame = customtkinter.CTkFrame(log_in_frame,width=825,height=325,fg_color="#D9D9D9",corner_radius=25)
login_clock_frame.place(x=260,y=470)

login_clock_label = customtkinter.CTkLabel(login_clock_frame,text="12:00 AM",width=725,height=195,text_color="white",fg_color="#292D32",corner_radius=25,font=("Quando",45))
login_clock_label.place(x=50,y=50)

login_no_account_label = customtkinter.CTkLabel(login_clock_frame,text="Don't have an account?",width=150,height=25,fg_color="transparent",text_color="black",corner_radius=25,font=("Quando",20))
login_no_account_label.place(x=140,y=255)

login_register_button = customtkinter.CTkButton(login_clock_frame,text="Register",command=to_sign_in,corner_radius=450,width=150,height=25,fg_color="#9A9A9A",text_color="black",font=("Quando",20))
login_register_button.place(x=440,y=255)
log_in_clock()

#\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\MAIN CONTENT FRAMES\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\#

# Opened Book Frame

opened_book_picture_frame = customtkinter.CTkFrame(open_book_frame, width=250,height=300,fg_color="#414240",corner_radius=0)
opened_book_picture_frame.place(x=140,y=20)

opened_book_author_title_frame = customtkinter.CTkFrame(open_book_frame, width=720,height=90,fg_color="#B88B68",corner_radius=0)
opened_book_author_title_frame.place(x=430, y=20)

opened_book_other_text_frame = customtkinter.CTkFrame(open_book_frame, width=720,height=170,fg_color="#614D40",corner_radius=0)
opened_book_other_text_frame.place(x=430, y=110)

opened_book_synopsis_title = customtkinter.CTkFrame(open_book_frame, width=1005,height=80,fg_color="#614D40",corner_radius=0)
opened_book_synopsis_title.place(x=140, y=350)

opened_book_synopsis_textbody = customtkinter.CTkFrame(open_book_frame, width=1005,height=800,fg_color="#B88B68",corner_radius=0)
opened_book_synopsis_textbody.place(x=140, y=420)

book_status_combobox = customtkinter.CTkComboBox(open_book_frame,border_color="#614D40", fg_color="#B88B68", dropdown_fg_color="#B88B68", values=["Book Status", "Reading", "Plan to read", "Completed", "Dropped"]) 
book_status_combobox.place(x=430, y=290)

book_title_label = customtkinter.CTkLabel(opened_book_author_title_frame, width=720, height=40, font=("Quando", 25),  text_color="white", text="Book title goes here.")
book_title_label.place(x=10, y=10)

book_author_label = customtkinter.CTkLabel(opened_book_author_title_frame, width=720, height=20, font=("Quando", 20),  text_color="white", text="Author name goes here.")
book_author_label.place(x=10, y=60)

book_genre_title = customtkinter.CTkLabel(opened_book_other_text_frame, anchor=customtkinter.W, width=150, height=30, font=("Quando", 22), text_color="white", text="GENRE: ")
book_genre_title.place(x=10, y=20)

book_genre_placeholder = customtkinter.CTkLabel(opened_book_other_text_frame, anchor=customtkinter.W,width=570, height=30, font=("Quando", 22), text_color="white",  text="Genre1, Genre2, Genre3")
book_genre_placeholder.place(x=170, y=20)#

book_theme_title = customtkinter.CTkLabel(opened_book_other_text_frame, anchor=customtkinter.W, width=150, height=30, font=("Quando", 22), text_color="white",  text="THEME: ")
book_theme_title.place(x=10, y=55)

book_theme_placeholder = customtkinter.CTkLabel(opened_book_other_text_frame,anchor=customtkinter.W, width=570, height=30, font=("Quando", 22),  text_color="white", text="Theme1, Theme2, Theme3")
book_theme_placeholder.place(x=170, y=55)#

book_keywords_title = customtkinter.CTkLabel(opened_book_other_text_frame, anchor=customtkinter.W, width=200, height=30, font=("Quando", 22), text_color="white",  text="KEYWORDS: ")
book_keywords_title.place(x=10, y=90)

book_keywords_placeholder = customtkinter.CTkLabel(opened_book_other_text_frame,anchor=customtkinter.W, width=570, height=30, font=("Quando", 22), text_color="white",  text="Keyword1, Keyword2, Keyword3")
book_keywords_placeholder.place(x=170, y=90)#

book_rating_title = customtkinter.CTkLabel(opened_book_other_text_frame, anchor=customtkinter.W, width=150,  height=30, font=("Quando", 22), text_color="white",  text="RATING: ")
book_rating_title.place(x=10, y=125)

book_rating_placeholder = customtkinter.CTkLabel(opened_book_other_text_frame,anchor=customtkinter.W, width=570, height=30, font=("Quando", 22), text_color="white",  text="5/5")
book_rating_placeholder.place(x=170, y=125)#

book_synopsis_title = customtkinter.CTkLabel(opened_book_synopsis_title, anchor=customtkinter.W, width=1005, height=85, font=("Quando", 45),  text_color="white", text="SYNOPSIS")
book_synopsis_title.place(x=15, y=-10)#

book_synopsis_body = customtkinter.CTkLabel(opened_book_synopsis_textbody, width=1005, height=800, font=("Quando", 25), wraplength=980, justify="center", text_color="white", text="My name is Yoshikage Kira. I'm 33 years old. My house is in the northeast section of Morioh, where all the villas are, and I am not married. I work as an employee for the Kame Yu department stores, and I get home every day by 8 PM at the latest. I don't smoke, but I occasionally drink. I'm in bed by 11 PM, and make sure I get eight hours of sleep, no matter what. After having a glass of warm milk and doing about twenty minutes of stretches before going to bed, I usually have no problems sleeping until morning. Just like a baby, I wake up without any fatigue or stress in the morning. I was told there were no issues at my last check-up. I'm trying to explain that I'm a person who wishes to live a very quiet life. I take care not to trouble myself with any enemies, like winning and losing, that would cause me to lose sleep at night. That is how I deal with society, and I know that is what brings me happiness. Although, if I were to fight I wouldn't lose to anyone.")
book_synopsis_body.place(x=0, y=-160)#

# Home Content
home_search_entry_button_img = customtkinter.CTkImage((Image.open("icons\\searchicon.png")),size=(30,30))
home_bookmark_frame = customtkinter.CTkFrame(home_frame, width=453, height=120, fg_color="#606060", corner_radius=0)
home_bookmark_frame.place(x = 570, y = 80)

home_user_frame = customtkinter.CTkFrame(home_frame, width=310, height=120, fg_color="#B88B68", corner_radius=0)
home_user_frame.place(x = 358, y = 80)

home_user_label = customtkinter.CTkLabel(home_user_frame,anchor=customtkinter.W,font=("Quando",25),text_color="white",width=100,height=30,fg_color="transparent",corner_radius=0, text="my Username")
home_user_label.place(x = 75, y = 30)

home_date_label = customtkinter.CTkLabel(home_user_frame,anchor=customtkinter.W,font=("Quando",15),text_color="white",width=100,height=10,fg_color="transparent",corner_radius=0, text="Date Joined: " )
home_date_label.place(x = 75, y = 70)

home_bookmark_label = customtkinter.CTkLabel(home_bookmark_frame,anchor=customtkinter.W,font=("Quando",20),text_color="white",width=100,height=30,fg_color="transparent",corner_radius=0,text="Bookmark: " )
home_bookmark_label.place(x = 136, y = 40)

home_upload_label = customtkinter.CTkLabel(home_bookmark_frame,anchor=customtkinter.W,font=("Quando",20),text_color="white",width=100,height=30,fg_color="transparent",corner_radius=0,text="Uploaded: " )
home_upload_label.place(x = 286, y = 40)

home_upload_frame_label_img = customtkinter.CTkImage((Image.open("icons\\camera.png")), size=(25,25))
home_upload_frame_img = customtkinter.CTkButton(home_frame, text="", image=home_upload_frame_label_img, width=120,height=120, fg_color="#614D40", corner_radius=0)
home_upload_frame_img.place(x = 270, y = 80)

home_lithit_frame_outer = customtkinter.CTkFrame(home_frame,width=485,height=510,fg_color="transparent",corner_radius=0)
home_lithit_frame_outer.place(x = 176, y = 260)

home_lithit_frame_inner = customtkinter.CTkScrollableFrame(home_lithit_frame_outer,width=455,height=450,fg_color="#B88B68",corner_radius=0)
home_lithit_frame_inner.place(x = 10, y = 40)

home_lithit_button = customtkinter.CTkButton(home_lithit_frame_outer,width=150,height=30,font=("Quando",20), command=lambda: show_main_frame_content("lithit"),fg_color="#737373",corner_radius=0,text="Literary Hits")
home_lithit_button.place(x = 10, y = 10)

home_arrowlit_button_img = customtkinter.CTkImage((Image.open("icons\\kanan.png")))
home_arrowlit_button_label = customtkinter.CTkButton(home_frame,command=lambda: show_main_frame_content("lithit"), image=home_arrowlit_button_img, width=25, height=25, fg_color="transparent", corner_radius=0, text="")
home_arrowlit_button_label.place(x = 635, y = 270)

home_lastread_frame_outer = customtkinter.CTkFrame(home_frame,width=485,height=490,fg_color="transparent",corner_radius=0)
home_lastread_frame_outer.place(x = 670, y = 260)

home_lastread_frame_inner = customtkinter.CTkScrollableFrame(home_lastread_frame_outer,width=465,height=440,fg_color="#B88B68",corner_radius=0)
home_lastread_frame_inner.place(x = 10, y = 40)

home_lastread_button = customtkinter.CTkButton(home_lastread_frame_outer,width=150,height=30,font=("Quando",20),fg_color="#737373",corner_radius=0,text="Last Read",command=open_last_read)
home_lastread_button.place(x = 10, y = 10)

home_arrowlast_button_img = customtkinter.CTkImage((Image.open("icons\\kanan.png")))
home_arrowlast_button_label = customtkinter.CTkButton(home_frame, image=home_arrowlast_button_img, width=25, height=25, fg_color="transparent", corner_radius=0, text="",command=open_last_read)
home_arrowlast_button_label.place(x = 1130, y = 270)

# update_lastread_home_stack()

# Lit Hit Content
lithit_main_frame = customtkinter.CTkFrame(lithit_frame, width=1360, height=765,fg_color="#F2E8BD")
lithit_main_frame.place(x=0, y=0)

lithit_label = customtkinter.CTkLabel(lithit_main_frame,font=("Quando",45),width=650,height=50,fg_color="transparent",text="Literature Hits")
lithit_label.place(x = 380, y = 20)

lithit_firstShelf_frame = customtkinter.CTkScrollableFrame(lithit_main_frame, width=1000 , height=635, fg_color="#B88B68")
lithit_firstShelf_frame.place(x=170, y=120)

display_lit_hit_data()

# Last Read Content
lRead_search_entry_frame = customtkinter.CTkFrame(lastread_frame,width=650,height=50,fg_color="transparent",corner_radius=45)
lRead_search_entry_frame.place(x = 440, y = 10)

lRead_search_entry = customtkinter.CTkLabel(lRead_search_entry_frame,width=550,height=50,fg_color="transparent",text="Last Read",font=("Quando",35))
lRead_search_entry.place(x = 5,y=5)

lRead_firstShelf_frame = customtkinter.CTkScrollableFrame(lastread_frame, width=1165 , height=650, fg_color="#B88B68")
lRead_firstShelf_frame.place(x=100, y=70)

# Search Content
search_advanceSearch_frame = customtkinter.CTkFrame(search_frame,width=1360,height=765,fg_color="#F2E8BD",corner_radius=0)
search_advanceSearch_frame.place(x = 0, y = 0)

search_label = customtkinter.CTkLabel(search_advanceSearch_frame,width=100,height=40,font=("Quando",35),fg_color="transparent",corner_radius=0,text="Library")
search_label.place(x = 165, y = 25)

search_inner_frame = customtkinter.CTkFrame(search_advanceSearch_frame,width=1000,height=850,fg_color="#433F3D",corner_radius=0)
search_inner_frame.place(x = 170, y = 70)

search_bookshelf_frame = customtkinter.CTkScrollableFrame(search_inner_frame,width=900,height=600,fg_color="#B88B68")
search_bookshelf_frame.place(x = 40, y = 78)

book_stack()

# Library Content

library_main_frame = customtkinter.CTkFrame(library_frame,width=1360,height=765,fg_color="#F2E8BD",corner_radius=0)
library_main_frame.place(x = 0, y = 0)

library_inner_frame = customtkinter.CTkFrame(library_main_frame,width=1000,height=850,fg_color="#433F3D",corner_radius=0)
library_inner_frame.place(x = 170, y = 70)

library_bookshelf_frame = customtkinter.CTkFrame(library_inner_frame,width=900,height=840,fg_color="transparent")
library_bookshelf_frame.place(x = 40, y = 78)

# book_stack()

library_search_entry_frame = customtkinter.CTkFrame(library_main_frame,width=650,height=50,fg_color="transparent",corner_radius=45)
library_search_entry_frame.place(x = 725,y=20)

search_var = tk.StringVar()
library_search_entry = customtkinter.CTkEntry(library_search_entry_frame,textvariable=search_var,width=450,height=45,fg_color="white",corner_radius=45,placeholder_text="Search Titles...",font=("Quando",20))
library_search_entry.place(x = 0,y=0)

library_search_entry_button_img = customtkinter.CTkImage((Image.open("icons\\searchicon.png")),size=(55,55))
library_search_entry_button = customtkinter.CTkButton(library_search_entry,text="",image=home_search_entry_button_img,width=10,height=10,fg_color="transparent")
library_search_entry_button.place(x = 390 , y=2.5)

library_label = customtkinter.CTkLabel(library_main_frame,width=100,height=40,font=("Quando",35),fg_color="transparent",corner_radius=0,text="Browse")
library_label.place(x = 165, y = 25)

# Addbooks Content
add_books_main_frame = customtkinter.CTkFrame(add_books_frame,width=1260,height=760,fg_color="transparent",corner_radius=0)
add_books_main_frame.place(x=60,y=0)

addbooks_upload_cover_frame = customtkinter.CTkFrame(add_books_main_frame,width=250,height=360,fg_color="#414240",corner_radius=0)
addbooks_upload_cover_frame.place(x=140,y=10)

addbooks_upload_cover_button_img = customtkinter.CTkImage(Image.open('icons\\coverbookupload.png'),size=(25,25))
addbooks_upload_cover_button = customtkinter.CTkButton(add_books_main_frame,image=addbooks_upload_cover_button_img,text_color="black",font=("Quando",15),text="Upload Cover",width=250,height=50,fg_color="#C8C8C8",corner_radius=25)
addbooks_upload_cover_button.place(x=140,y=380)
addbooks_upload_cover_button.configure(command=upload_cover)

addbooks_book_detail_input_frame = customtkinter.CTkFrame(add_books_main_frame,width=715,height=420,fg_color="#808080",corner_radius=0)
addbooks_book_detail_input_frame.place(x=405,y=10)

addbooks_book_detail_input_title_entry = customtkinter.CTkEntry(addbooks_book_detail_input_frame,font=("Quando",25),width=695,height=55,fg_color="#F1F1F1",corner_radius=25,placeholder_text="Input Book Title",placeholder_text_color="gray")
addbooks_book_detail_input_title_entry.place(x=10,y=10)

addbooks_book_detail_input_author_name_entry = customtkinter.CTkEntry(addbooks_book_detail_input_frame,font=("Quando",15),width=525,height=45,fg_color="#F1F1F1",corner_radius=25,placeholder_text="Input Author(s) Name",placeholder_text_color="gray")
addbooks_book_detail_input_author_name_entry.place(x=10,y=75)

addbooks_book_detail_input_publish_date_entry = customtkinter.CTkEntry(addbooks_book_detail_input_frame,font=("Quando",12),width=165,height=45,fg_color="#F1F1F1",corner_radius=25,placeholder_text="Publish Date",placeholder_text_color="gray")
addbooks_book_detail_input_publish_date_entry.place(x=540,y=75)

addbooks_book_detail_input_inner_frame = customtkinter.CTkFrame(addbooks_book_detail_input_frame,width=715,height=290,fg_color="#C8C8C8",corner_radius=0)
addbooks_book_detail_input_inner_frame.place(x=0,y=130)

addbooks_book_detail_input_genre_label = customtkinter.CTkLabel(addbooks_book_detail_input_inner_frame,text="Genre:",font=("Quando",20),width=75,height=55,fg_color="transparent")
addbooks_book_detail_input_genre_label.place(x=10,y=10)

addbooks_book_detail_input_theme_label = customtkinter.CTkLabel(addbooks_book_detail_input_inner_frame,text="Theme:",font=("Quando",20),width=75,height=55,fg_color="transparent")
addbooks_book_detail_input_theme_label.place(x=10,y=65)

addbooks_book_detail_input_keywords_label = customtkinter.CTkLabel(addbooks_book_detail_input_inner_frame,text="Keywords:",font=("Quando",20),width=75,height=55,fg_color="transparent")
addbooks_book_detail_input_keywords_label.place(x=10,y=120)

addbooks_book_detail_input_language_label = customtkinter.CTkLabel(addbooks_book_detail_input_inner_frame,text="Language:",font=("Quando",20),width=75,height=55,fg_color="transparent")
addbooks_book_detail_input_language_label.place(x=10,y=175)

addbooks_book_detail_input_rating_label = customtkinter.CTkLabel(addbooks_book_detail_input_inner_frame,text="Rating:",font=("Quando",20),width=75,height=55,fg_color="transparent")
addbooks_book_detail_input_rating_label.place(x=10,y=235)

addbooks_book_detail_input_genre_entry = customtkinter.CTkEntry(addbooks_book_detail_input_inner_frame,font=("Quando",12),width=565,height=45,fg_color="#F1F1F1",corner_radius=25)
addbooks_book_detail_input_genre_entry.place(x=135,y=10)

addbooks_book_detail_input_theme_entry = customtkinter.CTkEntry(addbooks_book_detail_input_inner_frame,font=("Quando",12),width=565,height=45,fg_color="#F1F1F1",corner_radius=25)
addbooks_book_detail_input_theme_entry.place(x=135,y=65)

addbooks_book_detail_input_keywords_entry = customtkinter.CTkEntry(addbooks_book_detail_input_inner_frame,font=("Quando",12),width=565,height=45,fg_color="#F1F1F1",corner_radius=25)
addbooks_book_detail_input_keywords_entry.place(x=135,y=120)

addbooks_book_detail_input_language_entry = customtkinter.CTkEntry(addbooks_book_detail_input_inner_frame,font=("Quando",12),width=565,height=45,fg_color="#F1F1F1",corner_radius=25)
addbooks_book_detail_input_language_entry.place(x=135,y=175)

addbooks_book_detail_rating = customtkinter.StringVar(value="other")

my_rad1 = customtkinter.CTkRadioButton(addbooks_book_detail_input_inner_frame, text="1", value="1", variable=addbooks_book_detail_rating,)
my_rad1.place(x=155,y=245)

my_rad2 = customtkinter.CTkRadioButton(addbooks_book_detail_input_inner_frame, text="2", value="2", variable=addbooks_book_detail_rating,)
my_rad2.place(x=215,y=245)

my_rad3 = customtkinter.CTkRadioButton(addbooks_book_detail_input_inner_frame, text="3", value="3", variable=addbooks_book_detail_rating,)
my_rad3.place(x=265,y=245)

my_rad4 = customtkinter.CTkRadioButton(addbooks_book_detail_input_inner_frame, text="4", value="4", variable=addbooks_book_detail_rating,)
my_rad4.place(x=315,y=245)

my_rad5 = customtkinter.CTkRadioButton(addbooks_book_detail_input_inner_frame, text="5", value="5", variable=addbooks_book_detail_rating,)
my_rad5.place(x=365,y=245)

addbooks_synopsis_frame = customtkinter.CTkFrame(add_books_main_frame,width=980,height=450,fg_color="#808080",corner_radius=0)
addbooks_synopsis_frame.place(x=140,y=445)

addbooks_synopsis_frame_label = customtkinter.CTkLabel(addbooks_synopsis_frame,anchor="w",text="Synopsis",width=980,height=75,fg_color="transparent",corner_radius=0,text_color="white",font=("Quando",35))
addbooks_synopsis_frame_label.place(x=20,y=0)

addbooks_synopsis_textbox = customtkinter.CTkTextbox(addbooks_synopsis_frame,width=980,height=235,fg_color="#D9D9D9",corner_radius=0,font=("Quando",25))
addbooks_synopsis_textbox.place(x=0,y=85)

addbooks_upload_button_img = customtkinter.CTkImage((Image.open("icons\\uploadpic.png")),size=(35,35))
addbooks_upload_button = customtkinter.CTkButton(add_books_main_frame,image=addbooks_upload_button_img,fg_color="#B88B68",text="",width=50,height=75,corner_radius=0,hover_color="#737373",command=add_book)
addbooks_upload_button.place(x =1120, y = 685)

# Title and Icon Frame Content
button_stackread_img = customtkinter.CTkImage((Image.open("icons\\stackreadicon.png")),size=(45,45))
button_stackread = customtkinter.CTkButton(icon_frame,image=button_stackread_img,text="",width=75,height=75,command=side_panel_switch,corner_radius=0,fg_color="#C8DF8C",hover_color="#737373",hover=False)
button_stackread.place(x = 0, y = 10)

label_title_img = customtkinter.CTkImage(Image.open('icons\\logo.png'),size=(145,25))
label_title = customtkinter.CTkLabel(title_frame,image=label_title_img,anchor=customtkinter.W, text="",width=150,height=75,font=("Quando",20),fg_color="transparent",text_color="black")
label_title.place(x = 0, y = 10)

# Buttons for sidebar
button_home_img = customtkinter.CTkImage((Image.open("icons\\homeunclicked.png")),size=(45,45))
button_home = customtkinter.CTkButton(icon_frame, image=button_home_img,text="", command=lambda: show_main_frame_content("Home"),width=75,height=75,corner_radius=0,fg_color="#C8DF8C",hover=False)
button_home.place(x = 0 , y = 100)

button_lithit_img = customtkinter.CTkImage((Image.open("icons\\lithitsunclicked.png")),size=(35,45))
button_lithit = customtkinter.CTkButton(icon_frame,image=button_lithit_img, text="", command=lambda: show_main_frame_content("lithit"),width=75,height=75,corner_radius=0,font=("Quando",16),fg_color="#C8DF8C",hover=False)
button_lithit.place(x = 0 , y = 180)

button_search_img = customtkinter.CTkImage((Image.open("icons\\searchunclicked.png")),size=(45,45))
button_search = customtkinter.CTkButton(icon_frame,image=button_search_img, text="", command=lambda: show_main_frame_content("Search"),width=75,height=75,corner_radius=0,font=("Quando",16),fg_color="#C8DF8C",hover=False)
button_search.place(x = 0 , y = 260)

button_library_img = customtkinter.CTkImage((Image.open("icons\\libraryunclicked.png")),size=(45,45))
button_library = customtkinter.CTkButton(icon_frame,image=button_library_img, text="", command=lambda: show_main_frame_content("Library"),width=75,height=75,corner_radius=0,font=("Quando",16),fg_color="#C8DF8C",hover=False)
button_library.place(x = 0 , y = 340)

button_addbooks_img= customtkinter.CTkImage((Image.open("icons\\disclickedaddbooks.png")),size=(45,45))
button_addbooks = customtkinter.CTkButton(icon_frame,image=button_addbooks_img,text="",command=lambda: show_main_frame_content("Add"),width=75,height=75,corner_radius=0,font=("Quando",16),fg_color="#C8DF8C",hover=False)
button_addbooks.place(x = 0, y = 420)

button_log_out_img= customtkinter.CTkImage((Image.open("icons\\logout.png")),size=(35,35))
button_log_out = customtkinter.CTkButton(icon_frame,image=button_log_out_img, text="", command=lambda: logout(),width=75,height=75,corner_radius=0,font=("Quando",16),fg_color="#C8DF8C",hover=False,text_color="black")
button_log_out.place(x = 5, y = 675)

# Button Labels for Sidebar
button_label_home = customtkinter.CTkButton(title_frame,anchor="w", text="Home",width=150,height=75,corner_radius=0,fg_color="transparent",font=("Quando",15),text_color="black",hover=False)
button_label_home.place(x = 0 , y = 100)

button_label_lithit = customtkinter.CTkButton(title_frame,anchor="w", text="Literary Hits",width=150,height=75,corner_radius=0,fg_color="transparent",font=("Quando",15),text_color="black",hover=False)
button_label_lithit.place(x = 0 , y = 180)

button_label_search = customtkinter.CTkButton(title_frame,anchor="w", text="Library",width=150,height=75,corner_radius=0,fg_color="transparent",font=("Quando",15),text_color="black",hover=False)
button_label_search.place(x = 0 , y = 260)

button_label_library = customtkinter.CTkButton(title_frame,anchor="w", text="Browse",width=150,height=75,corner_radius=0,fg_color="transparent",font=("Quando",15),text_color="black",hover=False)
button_label_library.place(x = 0 , y = 340)

button_label_addbooks = customtkinter.CTkButton(title_frame,anchor="w", text="Add books",width=150,height=75,corner_radius=0,fg_color="transparent",font=("Quando",15),text_color="black",hover=False)
button_label_addbooks.place(x = 0, y = 420)

# button_label_log_out_img= customtkinter.CTkImage((Image.open("icons\disclickedaddbooks.png")),size=(50,50))
button_label_log_out = customtkinter.CTkButton(title_frame,anchor="w", text="Log out",width=150,height=75,corner_radius=0,fg_color="transparent",font=("Quando",15),text_color="black",hover=False)
button_label_log_out.place(x = 0, y = 675)

#\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\FOR SEARCH ALGORITHM\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\\#

# Node class for Linked List
class Node:
    def __init__(self, book_id, title, author, genre, theme, ownership, rating, keywords, language, awards, publish_date, image_path):
        self.book_id = book_id
        self.title = title
        self.author = author
        self.genre = genre
        self.theme = theme
        self.ownership = ownership
        self.rating = rating
        self.keywords = keywords
        self.language = language
        self.awards = awards
        self.publish_date = publish_date
        self.image_path = image_path
        self.next = None

# LinkedList class to manage the collection of books
class LinkedList:
    def __init__(self):
        self.head = None
        self.current = None

    def add_book(self, book_id, title, author, genre, theme, ownership, rating, keywords, language, awards, publish_date, image_path):
        new_book = Node(book_id, title, author, genre, theme, ownership, rating, keywords, language, awards, publish_date, image_path)
        if not self.head:
            self.head = new_book
            self.current = new_book
        else:
            current_book = self.head
            while current_book.next:
                current_book = current_book.next
            current_book.next = new_book

    def traverse(self):
        result = []
        current_book = self.head
        while current_book:
            result.append(current_book)
            current_book = current_book.next
        return result

    def binary_search(self, target_id):
        current_book = self.head
        while current_book:
            if current_book.book_id == target_id:
                return current_book
            current_book = current_book.next
        return None

    def search_by_title(self, search_query):
        result = []
        current_book = self.head
        while current_book:
            if search_query in current_book.title.lower():
                result.append(current_book)
            current_book = current_book.next
        return result

    def search_by_author(self, search_query):
        result = []
        current_book = self.head
        while current_book:
            if search_query in current_book.author.lower():
                result.append(current_book)
            current_book = current_book.next
        return result

    def search_by_attribute(self, search_query, attribute):
        result = []
        current_book = self.head
        while current_book:
            attribute_value = getattr(current_book, attribute, None)
            if attribute_value and search_query in attribute_value.lower():
                result.append(current_book)
            current_book = current_book.next
        return result

# Load data from the Excel file
def load_data_from_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    linked_list = LinkedList()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Handle missing values by setting them to None
        while len(row) < 12:
            row += (None,)

        book_id, title, author, genre, theme, ownership, rating, keywords, language, awards, publish_date, image_path, *extra_values = row
        linked_list.add_book(book_id, title, author, genre, theme, ownership, rating, keywords, language, awards, publish_date, image_path)

    return linked_list

# Create the initial frame
def create_results_frame():
    global results_frame
    if results_frame:
        results_frame.destroy()

    # Create a new frame to display search results in a scrollable area
    results_frame = customtkinter.CTkFrame(outer_frame,fg_color="transparent")
    results_frame.pack()

def display_search_results(results):
    global results_frame
    create_results_frame()

    # Create a Frame inside the canvas to hold the buttons
    inner_frame = customtkinter.CTkScrollableFrame(results_frame,fg_color="transparent",width=900,height=500)
    inner_frame.pack()

    def display_book_details_in_button(book):
        msg = CTkMessagebox(title="Details", message = f"Book ID: {book.book_id}\nTitle: {book.title}\nAuthor: {book.author}\nGenre: {book.genre}\nTheme: {book.theme}\nRating: {book.ownership}\nKeywords: {book.rating}\nLanguage: {book.keywords}\nAwards: {book.language}\nPublished: {book.awards}",icon="info", option_1="Ok")
 
    # Add buttons for each result to the inner frame
    for book in results:
        button_text = f"{book.title} - {book.author}"
        button = customtkinter.CTkButton(inner_frame,anchor="center", text=button_text, command=lambda b=book: display_book_details_in_button(b),width=1200,height=125,hover_color="#3D291E",fg_color="#614D40",font=("Quando",20),corner_radius=15)
        button.pack(pady=5)

def on_search_var_change(*args):
    library_bookshelf_frame.configure(fg_color="#B88B68")
    search()

def search():
    search_query = search_var.get().strip().lower()
    if not search_query:
        display_search_results(linked_list.traverse())
    else:
        result = linked_list.search_by_title(search_query)
        if not result:
            result = linked_list.search_by_author(search_query)
            if not result:
                result = linked_list.search_by_attribute(search_query, "genre")
                if not result:
                    result = linked_list.search_by_attribute(search_query, "theme")
                    if not result:
                        result = linked_list.search_by_attribute(search_query, "ownership")
                        if not result:
                            result = linked_list.search_by_attribute(search_query, "rating")
                            if not result:
                                result = linked_list.search_by_attribute(search_query, "keywords")
                                if not result:
                                    result = linked_list.search_by_attribute(search_query, "language")
                                    if not result:
                                        result = linked_list.search_by_attribute(search_query, "awards")
                                        if not result:
                                            result = linked_list.search_by_attribute(search_query, "publish_date")

        if result:
            display_search_results(result)
        else:
            print()
            book_stack()

# Load data from the Excel file
file_path = "spreadsheets\\books.xlsx"  # Replace with the path to your Excel file
linked_list = load_data_from_excel(file_path)

# Set the default current book to Book ID 1
linked_list.current = linked_list.binary_search(1)

# Create search entry and button
search_var.trace_add("write", on_search_var_change)

result_label = customtkinter.CTkLabel(library_bookshelf_frame, text="")
result_label.pack(side=tk.TOP, padx=10, pady=10)

# Create an outer frame that won't be destroyed
outer_frame = customtkinter.CTkFrame(library_bookshelf_frame,fg_color="transparent")
outer_frame.pack(expand=True)

# Declare results_frame as a global variable
results_frame = None


# Main Loop
Main_app.mainloop()