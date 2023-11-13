import openpyxl
import tkinter as tk
from tkinter import ttk
from collections import deque

# Create a linked list to store bookId values
book_id_linked_list = deque()

# Create a stack to store the last ten values
last_ten_stack = []

# Replace 'your_excel_file.xlsx' with the path to your Excel file
excel_file = 'spreadsheets\\books.xlsx'

# Load the Excel file
workbook = openpyxl.load_workbook(excel_file)

# Choose a specific sheet within the Excel file, e.g., the first sheet
sheet = workbook.active

# Use a variable to skip the first row (headers)
skip_first_row = True

# Iterate through the rows in the sheet
for row in sheet.iter_rows(values_only=True):
    if skip_first_row:
        skip_first_row = False
        continue

    book_id = row[0]
    book_id_linked_list.append(book_id)

# Create a function to update the last ten values stack
def update_last_ten_stack():
    last_ten_stack.clear()
    for book_id in list(book_id_linked_list)[-3:]:
        # Find the corresponding row in the Excel file and append it to the stack
        for row in sheet.iter_rows(values_only=True):
            if row[0] == book_id:
                book_info = {
                    'book_id': row[0],
                    'title': row[1],
                    'author': row[2]
                }
                last_ten_stack.append(book_info)
    
    display_last_ten_stack_as_buttons()

# Create a function to display the last ten values stack as buttons
def display_last_ten_stack_as_buttons():
    # Create buttons for each book in the last ten values stack and add them to the frame with spacing
    for i, book_info in enumerate(last_ten_stack):
        book_button = tk.Button(
            button_frame,
            text=f"Book ID: {book_info['book_id']}\nTitle: {book_info['title']}\nAuthor: {book_info['author']}",
            width=30,
            height=3,
        )
        book_button.pack(pady=5)

    # Update the frame size
    button_frame.update_idletasks()

    # Configure the canvas to scroll the frame
    canvas.configure(scrollregion=canvas.bbox("all"))

# Create a Tkinter window
root = tk.Tk()
root.title("Stack Display")

# Create a canvas to hold the scrollable frame
canvas = tk.Canvas(root)
canvas.pack(side="top", fill="both", expand=True)

# Create a frame to hold the buttons
button_frame = tk.Frame(canvas)
canvas.create_window((0, 0), window=button_frame, anchor="nw")

# Create a button to trigger displaying the last ten values as buttons
display_button = tk.Button(root, text="Display Last Ten Values as Buttons", command=update_last_ten_stack)
display_button.pack()

root.mainloop()
